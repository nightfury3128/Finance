from flask import Flask, render_template, redirect, url_for, session, request
import pandas as pd
import os
import requests
import json
import openpyxl 
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Set a secret key for session management

API_KEY = '2EJ4D4oNPhYy4RoOPDZb2Qy9G1UfmI6c'  # Financial Modeling Prep API Key

def is_email_registered(email, filename='users_data.xlsx'):
    """
    Check if the email is already registered in the Excel file.
    """
    try:
        df = pd.read_excel(filename)
        if email in df['Email ID'].values:
            return True
        else:
            return False
    except FileNotFoundError:
        # File not found, so email can't be registered yet
        return False
    
def get_current_stock_price(symbol):
    url = f'https://financialmodelingprep.com/api/v3/quote/{symbol}?apikey={API_KEY}'
    try:
        response = requests.get(url)
        data = response.json()
        if data:
            return data[0]['price']  # Assuming the API returns the current price in 'price' field
    except Exception as e:
        print(f"Error fetching current stock price for {symbol}: {e}")
    return None

def search_symbol_by_company_name(company_name):
    url = f'https://financialmodelingprep.com/api/v3/search?query={company_name}&limit=1&apikey={API_KEY}'
    try:
        response = requests.get(url)
        data = response.json()
        if data:
            # Assuming the first result is the most relevant one
            return data[0]['symbol']
    except Exception as e:
        print(f"Error searching for company name {company_name}: {e}")
    return None

def validate_and_get_symbols(company_names):
    validated_symbols = []
    for name in company_names.split(','):
        name = name.strip()
        symbol = search_symbol_by_company_name(name)
        if symbol:
            validated_symbols.append(symbol)
        else:
            print(f"Company {name} not found or does not have a stock symbol.")
    return validated_symbols

def save_to_excel(email, name, dob, income, country, stocks, username, password, filename='users_data.xlsx'):
    # Serialize the stocks information into JSON
    stocks_json = json.dumps(stocks)

    data_dict = {
        'Email ID': email,
        'Name': name,
        'Date of Birth': dob,
        'Monthly Income': income,
        'Country': country,
        'Username': username,
        'Password': password,
        'Stocks Info': stocks_json  # Storing the serialized JSON
    }
    
    df = pd.DataFrame([data_dict])
    
    if not os.path.exists(filename):
        # If the file does not exist, create it and write the dataframe
        df.to_excel(filename, engine='openpyxl', index=False)
    else:
        # If the file exists, append the new data
        book = load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        
        startrow = writer.sheets['Sheet1'].max_row
        
        # Append without overwriting the existing sheet
        df.to_excel(writer, sheet_name='Sheet1', startrow=startrow, index=False, header=False)
        
        writer.save()
def validate_credentials(username, password, filename='users_data.xlsx'):
    try:
        df = pd.read_excel(filename)
        df['Username'] = df['Username'].astype(str)
        df['Password'] = df['Password'].astype(str)

        # Case-insensitive comparison
        user_row = df.loc[df['Username'].str.lower() == username.lower()]

        if not user_row.empty and user_row.iloc[0]['Password'] == password:
            return True
        else:
            return False
    except FileNotFoundError:
        return False

def save_user_data(email, username, stocks, shares, invested):
    wb = load_workbook('users_data.xlsx')
    users_sheet = wb['Users']
    stocks_sheet = wb['Stocks']

    # Add user to 'Users' sheet
    users_sheet.append([email, username])
    # Add stocks to 'Stocks' sheet
    for stock_name, share, invest in zip(stocks, shares, invested):
        stocks_sheet.append([username, stock_name, share, invest])

    wb.save('users_data.xlsx')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        # User details
        email = request.form['email']
        username = request.form['username']
        # Other user details processing

        # Process stocks data from form
        stocks = request.form.getlist('stocks[name]')
        shares = request.form.getlist('stocks[shares]')
        invested = request.form.getlist('stocks[invested]')

        # Save user and stocks data to Excel
        save_user_data(email, username, stocks, shares, invested)

        # Redirect to profile or another page after registration
        return redirect(url_for('profile'))
    return render_template('register.html')

@app.route('/signin', methods=['GET', 'POST'])
def signin():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if validate_credentials(username, password):
            session['username'] = username  # Set session
            return redirect(url_for('profile'))  # Redirect to a profile route
        else:
            return 'Invalid Username or Password!'
    return render_template('signin.html')

# Define a new route for the profile page
@app.route('/profile')
@app.route('/profile')
@app.route('/profile')
def profile():
    if 'username' in session:
        username = session['username']
        # Fetch user data
        df_users = pd.read_excel('users_data.xlsx', sheet_name='Users')
        user_data = df_users[df_users['Username'].str.lower() == username.lower()].iloc[0]
        monthly_income = user_data['Monthly Income']

        # Assuming stocks information is stored in a separate sheet or structured format
        df_stocks = pd.read_excel('users_data.xlsx', sheet_name='Stocks')
        user_stocks = df_stocks[df_stocks['Username'].str.lower() == username.lower()]

        total_stock_value = 0
        total_profit = 0
        stocks_info = []

        for _, row in user_stocks.iterrows():
            symbol = row['StockName']
            num_shares = row['Shares']
            purchase_price = row['InvestedAmount'] / num_shares if num_shares else 0
            current_price = get_current_stock_price(symbol)

            if current_price is not None:
                stock_value = current_price * num_shares
                profit = (current_price - purchase_price) * num_shares

                total_stock_value += stock_value
                total_profit += profit

                stocks_info.append({
                    'symbol': symbol,
                    'num_shares': num_shares,
                    'purchase_price': purchase_price,
                    'current_price': current_price,
                    'profit': profit
                })

        net_worth = total_stock_value + monthly_income  # Simplified net worth calculation

        return render_template('profile.html', username=username, monthly_income=monthly_income, total_stock_value=total_stock_value, net_worth=net_worth, stocks_info=stocks_info, total_profit=total_profit)
    else:
        return redirect(url_for('signin'))

@app.route('/signout')
def sign_out():
    session.clear()  # Clear the user's session
    return redirect(url_for('signin'))

@app.route('/')
def home():
    return redirect(url_for('signin'))
if __name__ == '__main__':
    app.run(debug=True)